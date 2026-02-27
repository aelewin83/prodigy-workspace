from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app.models.enums import ListingSourceType
from app.services.comps import build_normalized_row


def _to_float(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_csv(file_path: str) -> tuple[list, dict]:
    rows = []
    dropped = 0
    with open(file_path, newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for raw in reader:
            address = raw.get("address") or raw.get("Address")
            if not address:
                dropped += 1
                continue
            date_val = raw.get("date_observed") or raw.get("Date Rented")
            date_observed = None
            if date_val:
                try:
                    date_observed = datetime.fromisoformat(str(date_val)).date()
                except ValueError:
                    date_observed = None

            rows.append(
                build_normalized_row(
                    address=address,
                    unit=raw.get("unit") or raw.get("Unit"),
                    beds=_to_float(raw.get("beds") or raw.get("Beds")),
                    baths=_to_float(raw.get("baths") or raw.get("Baths")),
                    rent=_to_float(raw.get("rent") or raw.get("Rent")),
                    gross_rent=_to_float(raw.get("gross_rent") or raw.get("Gross Rent")),
                    date_observed=date_observed,
                    link=raw.get("link") or raw.get("Link"),
                    notes=raw.get("notes") or raw.get("Notes"),
                    source_type=ListingSourceType.PRIVATE_FILE,
                    source_ref=file_path,
                    confidence_score=0.95,
                )
            )

    report = {"type": "csv", "rows_parsed": len(rows), "rows_dropped": dropped, "unmapped_columns": []}
    return rows, report


def parse_xlsx(file_path: str) -> tuple[list, dict]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    header = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    lookup = {name: idx for idx, name in enumerate(header)}
    rows = []
    dropped = 0

    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        address = row_vals[lookup.get("Address", -1)] if "Address" in lookup else None
        if not address:
            dropped += 1
            continue
        date_val = row_vals[lookup.get("Date Rented", -1)] if "Date Rented" in lookup else None
        date_observed = date_val.date() if hasattr(date_val, "date") else None

        rows.append(
            build_normalized_row(
                address=str(address),
                unit=str(row_vals[lookup.get("Unit", -1)]) if "Unit" in lookup and row_vals[lookup.get("Unit", -1)] else None,
                beds=_to_float(row_vals[lookup.get("Beds", -1)] if "Beds" in lookup else None),
                baths=_to_float(row_vals[lookup.get("Baths", -1)] if "Baths" in lookup else None),
                rent=_to_float(row_vals[lookup.get("Rent", -1)] if "Rent" in lookup else None),
                gross_rent=_to_float(row_vals[lookup.get("Gross Rent", -1)] if "Gross Rent" in lookup else None),
                date_observed=date_observed,
                link=str(row_vals[lookup.get("Link", -1)]) if "Link" in lookup and row_vals[lookup.get("Link", -1)] else None,
                notes=str(row_vals[lookup.get("Notes", -1)]) if "Notes" in lookup and row_vals[lookup.get("Notes", -1)] else None,
                source_type=ListingSourceType.PRIVATE_FILE,
                source_ref=file_path,
                confidence_score=0.9,
            )
        )

    report = {"type": "xlsx", "rows_parsed": len(rows), "rows_dropped": dropped, "unmapped_columns": []}
    return rows, report


def parse_pdf(file_path: str) -> tuple[list, dict]:
    # Deterministic fail-soft placeholder for PDF table parser integration.
    exists = Path(file_path).exists()
    report = {
        "type": "pdf",
        "rows_parsed": 0,
        "rows_dropped": 0,
        "unmapped_columns": [],
        "note": "PDF table extraction adapter not configured yet" if exists else "file not found",
    }
    return [], report
