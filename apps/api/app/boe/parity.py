from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

from app.boe.engine import BOEInput, calculate_boe, serialize_output


@dataclass(frozen=True)
class Tolerance:
    pct_abs: float = 0.001
    dollars_abs: float = 5.0


OUTPUT_PERCENT_KEYS = {
    "market_cap_rate",
    "asking_cap_rate",
    "analysis_cap_rate",
    "y1_exit_cap_rate",
    "y1_expense_ratio",
    "y1_cash_on_cash",
    "y1_yield_on_cost_unlevered",
}

OUTPUT_DOLLAR_KEYS = {
    "seller_noi_from_om",
    "residual_sale_at_exit_cap",
    "profit_potential",
    "max_price_at_yoc",
    "max_price_at_capex_multiple",
    "max_price_at_coc_threshold",
    "boe_max_bid",
    "delta_vs_asking",
    "deposit_amount",
}


def load_fixture(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def compare_case(case: dict, tolerance: Tolerance = Tolerance()) -> list[str]:
    errs: list[str] = []
    output, tests, decision = calculate_boe(BOEInput(**case["inputs"]))
    actual_outputs = serialize_output(output)
    expected = case["expected"]

    for key, exp in expected["outputs"].items():
        act = actual_outputs.get(key)
        if exp is None or act is None:
            if exp != act:
                errs.append(f"{case['name']} output {key}: expected {exp} got {act}")
            continue

        if isinstance(exp, str) or isinstance(act, str):
            if exp != act:
                errs.append(f"{case['name']} output {key}: expected {exp} got {act}")
            continue

        diff = abs(float(act) - float(exp))
        if key in OUTPUT_PERCENT_KEYS and diff > tolerance.pct_abs:
            errs.append(f"{case['name']} output {key}: diff {diff} exceeds {tolerance.pct_abs}")
        elif key in OUTPUT_DOLLAR_KEYS and diff > tolerance.dollars_abs:
            errs.append(f"{case['name']} output {key}: diff {diff} exceeds {tolerance.dollars_abs}")
        elif key not in OUTPUT_PERCENT_KEYS and key not in OUTPUT_DOLLAR_KEYS and diff > tolerance.pct_abs:
            errs.append(f"{case['name']} output {key}: diff {diff} exceeds {tolerance.pct_abs}")

    expected_tests = expected["tests"]
    actual_tests = {t.key: t.result.value for t in tests}
    for key, exp in expected_tests.items():
        act = actual_tests.get(key)
        if act != exp:
            errs.append(f"{case['name']} test {key}: expected {exp} got {act}")

    if decision.hard_veto_ok != expected["decision"]["hard_veto_ok"]:
        errs.append(
            f"{case['name']} hard_veto_ok expected {expected['decision']['hard_veto_ok']} got {decision.hard_veto_ok}"
        )
    if decision.pass_count != expected["decision"]["pass_count"]:
        errs.append(
            f"{case['name']} pass_count expected {expected['decision']['pass_count']} got {decision.pass_count}"
        )
    if decision.advance != expected["decision"]["advance"]:
        errs.append(f"{case['name']} advance expected {expected['decision']['advance']} got {decision.advance}")

    if output.binding_constraint != expected["binding_constraint"]:
        errs.append(
            f"{case['name']} binding_constraint expected {expected['binding_constraint']} got {output.binding_constraint}"
        )

    return errs


def run_fixture_parity(fixtures_dir: Path) -> list[str]:
    errs: list[str] = []
    for path in sorted(fixtures_dir.glob("*.json")):
        case = load_fixture(path)
        if "expected" not in case:
            continue
        errs.extend(compare_case(case))
    return errs


def workbook_available(path: str | Path) -> bool:
    return Path(path).exists()


def run_workbook_snapshot_parity(
    workbook_path: Path,
    cell_map: dict[str, dict[str, str]],
    case_inputs: dict,
) -> dict[str, float | str | None]:
    """
    Optional parity helper for environments where the Excel template is mounted.
    Uses openpyxl to write inputs and read output snapshot values from mapped cells.
    """
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, data_only=False)
    ws = wb.active
    for key, cell in cell_map.get("inputs", {}).items():
        if key in case_inputs:
            ws[cell] = case_inputs[key]
    # openpyxl cannot recalculate formulas; this returns cached cell values from workbook snapshots.
    wb_data = load_workbook(workbook_path, data_only=True)
    ws_data = wb_data.active
    out: dict[str, float | str | None] = {}
    for key, cell in cell_map.get("outputs", {}).items():
        out[key] = ws_data[cell].value
    return out
